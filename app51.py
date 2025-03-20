from dotenv import load_dotenv
load_dotenv()

import os
import json  # /income_detail_apply で使用
import matplotlib
matplotlib.use('Agg')  # 非対話型バックエンドを使用

import io
import pandas as pd
import matplotlib.pyplot as plt
from flask import Flask, render_template, request, redirect, flash, url_for, session
import base64
from matplotlib import font_manager
import webbrowser
from threading import Timer

app = Flask(__name__, template_folder='templates')
# 本番環境では環境変数からSECRET_KEYを取得
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24))

# 日本語フォントの設定
font_path = 'C:/Windows/Fonts/msgothic.ttc'
try:
    font_prop = font_manager.FontProperties(fname=font_path)
    plt.rcParams['font.family'] = font_prop.get_name()
except FileNotFoundError:
    plt.rcParams['font.family'] = 'sans-serif'

def get_tax_rate(income):
    if income <= 100:
        return 0
    elif income <= 150:
        return 13
    elif income <= 200:
        return 15
    elif income <= 250:
        return 18
    elif income <= 300:
        return 20
    elif income <= 400:
        return 21
    elif income <= 500:
        return 23
    elif income <= 600:
        return 24
    elif income <= 700:
        return 25
    elif income <= 800:
        return 26
    elif income <= 900:
        return 27
    elif income <= 1000:
        return 28
    elif income <= 1200:
        return 29
    elif income <= 1400:
        return 31
    elif income <= 1500:
        return 32
    elif income <= 2000:
        return 35
    elif income <= 3000:
        return 40
    elif income <= 5000:
        return 50
    elif income <= 10000:
        return 56
    else:
        return 60

def calculate_annual_repayment(principal, annual_interest_rate, years):
    if annual_interest_rate == 0:
        return principal / years
    r = annual_interest_rate / 100
    n = years
    payment = principal * r * (1 + r) ** n / ((1 + r) ** n - 1)
    return round(payment)

# working_style は、各年の働き方 ("company" または "self") の配列
def calculate_living_expenses(
    current_age, retirement_age, current_income, current_savings, peak_income, rent,
    car_start_age, car_price, car_interval, housing_status,
    retirement_money=0, loan_amount=0, loan_years=0, housing_expenses=0,
    special_expenses=[], special_incomes=[], housing_changes=[], insurances=[],
    employment_type=None, custom_incomes=None, working_style=None
):
    starting_age = 23
    person_ages = list(range(current_age, 91))
    
    # 年金計算の基本パラメータ
    starting_income = 300
    average_income = (starting_income + peak_income) / 2
    working_years = retirement_age - starting_age
    pension_amount = 80 if employment_type != 'company' else 80 + 0.22 * average_income * (working_years) / 40
    
    # working_style がある場合、会社員期間とそのピーク収入を再計算
    if working_style is not None:
        company_years = 0
        company_peak = 0
        var_limit = retirement_age - current_age
        for i in range(var_limit):
            if working_style[i] == "company":
                company_years += 1
                if custom_incomes is not None and i < len(custom_incomes):
                    income_val = custom_incomes[i]
                else:
                    income_val = current_income + (peak_income - current_income) * (i / var_limit)
                if income_val > company_peak:
                    company_peak = income_val
        if company_years > 0:
            pension_amount = 80 + 0.22 * company_peak * (company_years / 40)
        else:
            pension_amount = 80

    # 収入の計算
    income = []
    retirement_money_arr = []
    for age in person_ages:
        var_ret_money = 0
        if isinstance(retirement_money, list):
            for event in retirement_money:
                if event.get("age") == age:
                    var_ret_money += event.get("amount", 0)
        else:
            if age == retirement_age:
                var_ret_money = retirement_money
        if custom_incomes is not None and age <= retirement_age:
            index = age - current_age
            if index < len(custom_incomes):
                yearly_income = custom_incomes[index]
            else:
                yearly_income = custom_incomes[-1]
        else:
            if current_age <= age < retirement_age:
                yearly_income = round(current_income + (peak_income - current_income) / (retirement_age - current_age) * (age - current_age))
            else:
                if age == retirement_age:
                    base_income = peak_income
                elif age >= 65:
                    base_income = pension_amount
                else:
                    base_income = 0
                yearly_income = base_income
        yearly_income += var_ret_money
        income.append(round(yearly_income))
        retirement_money_arr.append(var_ret_money)
    
    # 税金・社会保険料の計算（退職金部分は一律10%）
    if working_style is not None:
        tax_social_insurance = []
        for i, x in enumerate(income):
            regular_income = x - retirement_money_arr[i]
            if working_style[i] == "company":
                reg_tax = round(regular_income * get_tax_rate(regular_income) / 100)
            else:
                reg_tax = 0
            ret_tax = round(retirement_money_arr[i] * 0.10)
            tax_social_insurance.append(reg_tax + ret_tax)
    else:
        if employment_type == 'company':
            tax_social_insurance = []
            for i, x in enumerate(income):
                regular_income = x - retirement_money_arr[i]
                reg_tax = round(regular_income * get_tax_rate(regular_income) / 100)
                ret_tax = round(retirement_money_arr[i] * 0.10)
                tax_social_insurance.append(reg_tax + ret_tax)
        else:
            tax_social_insurance = [round(r * 0.10) for r in retirement_money_arr]
    
    living_expenses = [216] * len(person_ages)
    if housing_status in ['賃貸', 'その他']:
        housing_expenses_list = [rent] * len(person_ages)
    elif housing_status == '住宅ローン':
        housing_expenses_list = []
        for age in person_ages:
            if loan_years > 0 and (age - current_age) < loan_years:
                annual_cost = loan_amount + housing_expenses
            else:
                annual_cost = housing_expenses
            housing_expenses_list.append(annual_cost)
    else:
        housing_expenses_list = [0] * len(person_ages)
    
    other_incomes = [0] * len(person_ages)
    other_expenses = [0] * len(person_ages)
    insurance_premiums = [0] * len(person_ages)
    
    for insurance in insurances:
        joining_age = insurance.get('joining_age', current_age) or current_age
        payment_end_age = insurance.get('payment_end_age')
        surrender_type = insurance.get('surrender_type')
        surrender_details = insurance.get('surrender_details', {})
        if joining_age < current_age:
            joining_age = current_age
        payment_end_age = payment_end_age or joining_age
        for age in person_ages:
            if joining_age <= age <= payment_end_age:
                idx = age - current_age
                if 0 <= idx < len(insurance_premiums):
                    insurance_premiums[idx] += insurance.get('premium', 0)
        if surrender_type == 'lump_sum':
            surrender_age = surrender_details.get('surrender_age')
            surrender_amount = surrender_details.get('surrender_amount', 0)
            if surrender_age and current_age <= surrender_age < 91:
                idx = surrender_age - current_age
                if 0 <= idx < len(person_ages):
                    other_incomes[idx] += surrender_amount
        elif surrender_type == 'annuity':
            annuity_start_age = surrender_details.get('annuity_start_age')
            annuity_end_age = surrender_details.get('annuity_end_age')
            annuity_amount = surrender_details.get('annuity_amount', 0)
            if annuity_start_age and annuity_end_age and annuity_start_age <= annuity_end_age:
                for age in person_ages:
                    if annuity_start_age <= age <= annuity_end_age:
                        idx = age - current_age
                        if 0 <= idx < len(other_incomes):
                            other_incomes[idx] += annuity_amount
        elif surrender_type == 'none':
            pass
    
    if housing_changes:
        sorted_changes = sorted(housing_changes, key=lambda x: x['change_age'])
        for change in housing_changes:
            change_age = change['change_age']
            new_type = change['housing_type']
            selling_profit = change.get('selling_profit', 0)
            misc_costs = change.get('misc_costs', 0)
            idx = change_age - current_age
            if idx < 0 or idx >= len(person_ages):
                continue
            if new_type in ['賃貸', 'その他']:
                annual_cost = change.get('annual_cost', 0)
                for i in range(idx, len(person_ages)):
                    housing_expenses_list[i] = annual_cost
            elif new_type == '購入':
                property_price = change.get('property_price', 0)
                down_payment = change.get('down_payment', 0)
                loan_years_change = change.get('loan_years', 0)
                loan_interest_rate = change.get('loan_interest_rate', 0)
                other_annual_cost = change.get('other_annual_cost', 0)
                loan_amount_calc = property_price - down_payment
                annual_repayment = calculate_annual_repayment(loan_amount_calc, loan_interest_rate, loan_years_change)
                total_annual_cost = annual_repayment + other_annual_cost
                for i, age in enumerate(person_ages):
                    if age >= change_age:
                        if (age - change_age) < loan_years_change:
                            housing_expenses_list[i] = total_annual_cost
                        else:
                            housing_expenses_list[i] = other_annual_cost
                if 0 <= idx < len(person_ages):
                    housing_expenses_list[idx] += down_payment
            else:
                continue
            if selling_profit > 0:
                if idx < len(person_ages):
                    other_incomes[idx] += selling_profit
            if misc_costs > 0:
                if idx < len(person_ages):
                    housing_expenses_list[idx] += misc_costs
    
    for expense in special_expenses:
        if expense['type'] == 'single':
            expense_age = expense['age']
            if current_age <= expense_age < 91:
                idx = expense_age - current_age
                if 0 <= idx < len(other_expenses):
                    other_expenses[idx] += expense['amount']
        elif expense['type'] == 'recurring':
            start_age = expense['start_age']
            end_age = expense['end_age']
            interval = expense['interval']
            amount = expense['amount']
            adjusted_start_age = max(start_age, current_age)
            for age in person_ages:
                if adjusted_start_age <= age <= end_age and (age - adjusted_start_age) % interval == 0:
                    idx = age - current_age
                    if 0 <= idx < len(other_expenses):
                        other_expenses[idx] += amount
    
    for income_item in special_incomes:
        if income_item['type'] == 'single':
            income_age = income_item['age']
            if current_age <= income_age < 91:
                idx = income_age - current_age
                if 0 <= idx < len(other_incomes):
                    other_incomes[idx] += income_item['amount']
        elif income_item['type'] == 'recurring':
            start_age = income_item['start_age']
            end_age = income_item['end_age']
            interval = income_item['interval']
            amount = income_item['amount']
            adjusted_start_age = max(start_age, current_age)
            for age in person_ages:
                if adjusted_start_age <= age <= end_age and (age - adjusted_start_age) % interval == 0:
                    idx = age - current_age
                    if 0 <= idx < len(other_incomes):
                        other_incomes[idx] += amount
    
    total_income = [income[i] + other_incomes[i] for i in range(len(person_ages))]
    
    net_savings = [
        total_income[i] - (living_expenses[i] + housing_expenses_list[i] + other_expenses[i] + tax_social_insurance[i] + insurance_premiums[i])
        for i in range(len(person_ages))
    ]
    assets = [current_savings]
    for i in range(1, len(net_savings)):
        assets.append(assets[-1] + net_savings[i])
    
    for _ in range(1000):
        if assets[-1] > 0:
            living_expenses = [x + 1 for x in living_expenses]
        elif assets[-1] < 0:
            living_expenses = [x - 1 for x in living_expenses]
        else:
            break
        net_savings = [
            total_income[i] - (living_expenses[i] + housing_expenses_list[i] + other_expenses[i] + tax_social_insurance[i] + insurance_premiums[i])
            for i in range(len(person_ages))
        ]
        assets = [current_savings]
        for i in range(1, len(net_savings)):
            assets.append(assets[-1] + net_savings[i])
        if abs(assets[-1]) < 1:
            assets[-1] = 0
            break
    
    peak_asset = max(assets)
    peak_age = person_ages[assets.index(peak_asset)]
    
    income = [int(round(x)) for x in income]
    tax_social_insurance = [int(round(x)) for x in tax_social_insurance]
    living_expenses = [int(round(x)) for x in living_expenses]
    housing_expenses_list = [int(round(x)) for x in housing_expenses_list]
    other_expenses = [int(round(x)) for x in other_expenses]
    other_incomes = [int(round(x)) for x in other_incomes]
    net_savings = [int(round(x)) for x in net_savings]
    assets = [int(round(x)) for x in assets]
    insurance_premiums = [int(round(x)) for x in insurance_premiums]
    
    return living_expenses[0], person_ages, assets, income, living_expenses, other_expenses, tax_social_insurance, other_incomes, housing_expenses_list, insurance_premiums, peak_age, peak_asset

@app.route('/')
def index():
    return redirect(url_for('step1'))

@app.route('/step1', methods=['GET', 'POST'])
def step1():
    if request.method == 'POST':
        try:
            current_age = int(request.form['age'])
            current_savings = int(request.form['savings'])
        except ValueError:
            flash("すべてのフィールドに正しい数値を入力してください。")
            return redirect(url_for('step1'))
        if not (0 <= current_age <= 100):
            flash("年齢は0歳から100歳の間で入力してください。")
            return redirect(url_for('step1'))
        if not (current_savings >= 0):
            flash("現在の貯蓄額は0以上で入力してください。")
            return redirect(url_for('step1'))
        session['age'] = current_age
        session['savings'] = current_savings
        return redirect(url_for('step2'))
    return render_template('step1.html')

@app.route('/step2', methods=['GET', 'POST'])
def step2():
    if 'age' not in session:
        flash("まず基本情報を入力してください。")
        return redirect(url_for('step1'))
    if request.method == 'POST':
        housing_status = request.form.get('housing_status')
        rent = request.form.get('rent', 0)
        loan_amount = request.form.get('loan_amount', 0)
        loan_years = request.form.get('loan_years', 0)
        housing_expenses = request.form.get('housing_expenses', 0)
        try:
            rent = int(rent) if rent else 0
            loan_amount = int(loan_amount) if loan_amount else 0
            loan_years = int(loan_years) if loan_years else 0
            housing_expenses = int(housing_expenses) if housing_expenses else 0
        except ValueError:
            flash("すべてのフィールドに正しい数値を入力してください。")
            return redirect(url_for('step2'))
        if housing_status not in ['賃貸', '住宅ローン', 'その他']:
            flash("有効な住まいの選択肢を選んでください。")
            return redirect(url_for('step2'))
        if housing_status in ['賃貸', 'その他']:
            if rent < 0:
                flash("年間住居費は0以上で入力してください。")
                return redirect(url_for('step2'))
        elif housing_status == '住宅ローン':
            if loan_amount < 0:
                flash("ローンの年間支払額は0以上で入力してください。")
                return redirect(url_for('step2'))
            if loan_years <= 0:
                flash("ローンの残り年数は1年以上で入力してください。")
                return redirect(url_for('step2'))
            if housing_expenses < 0:
                flash("年間住宅関連費用は0以上で入力してください。")
                return redirect(url_for('step2'))
        session['housing_status'] = housing_status
        session['rent'] = rent
        session['loan_amount'] = loan_amount
        session['loan_years'] = loan_years
        session['housing_expenses'] = housing_expenses
        return redirect(url_for('step3'))
    return render_template('step2.html')

@app.route('/step3', methods=['GET', 'POST'])
def step3():
    if 'age' not in session or 'savings' not in session:
        flash("まず基本情報を入力してください。")
        return redirect(url_for('step1'))
    if request.method == 'POST':
        try:
            employment_type = request.form['employment_type']
            if employment_type == "company":
                current_income = int(request.form['company_income'])
                retirement_age = int(request.form['company_retirement_age'])
                peak_income = int(request.form['company_peak_income'])
                if request.form.get('company_retirement_option') == 'yes':
                    ages_list = request.form.getlist('company_retirement_age[]')
                    amounts_list = request.form.getlist('company_retirement_amount[]')
                    var_has_value = any(a.strip() for a in ages_list)
                    if var_has_value:
                        new_retirement_money = []
                        for a, amt in zip(ages_list, amounts_list):
                            try:
                                a_val = int(a)
                                amt_val = int(amt)
                                new_retirement_money.append({"age": a_val, "amount": amt_val})
                            except ValueError:
                                continue
                        retirement_money = new_retirement_money
                    else:
                        retirement_money = session.get('retirement_money', [])
                else:
                    retirement_money = session.get('retirement_money', [])
            else:
                current_income = int(request.form['self_income'])
                retirement_age = int(request.form['self_retirement_age'])
                peak_income = int(request.form['self_peak_income'])
                if request.form.get('self_retirement_option') == 'yes':
                    ages_list = request.form.getlist('self_retirement_age[]')
                    amounts_list = request.form.getlist('self_retirement_amount[]')
                    var_has_value = any(a.strip() for a in ages_list)
                    if var_has_value:
                        new_retirement_money = []
                        for a, amt in zip(ages_list, amounts_list):
                            try:
                                a_val = int(a)
                                amt_val = int(amt)
                                new_retirement_money.append({"age": a_val, "amount": amt_val})
                            except ValueError:
                                continue
                        retirement_money = new_retirement_money
                    else:
                        retirement_money = session.get('retirement_money', [])
                else:
                    retirement_money = session.get('retirement_money', [])
        except ValueError:
            flash("すべてのフィールドに正しい数値を入力してください。")
            return redirect(url_for('step3'))
        if not (current_income > 0):
            flash("現在の年収（または年間手取り）は正の値で入力してください。")
            return redirect(url_for('step3'))
        if not (retirement_age > session['age'] and retirement_age <= 100):
            flash("退職年齢（または引退年齢）は現在の年齢より高く、100歳以下で入力してください。")
            return redirect(url_for('step3'))
        session['current_income'] = current_income
        session['retirement_age'] = retirement_age
        if 'custom_incomes' in session:
            session['peak_income'] = max(session['custom_incomes'])
        else:
            session['peak_income'] = peak_income
        session['retirement_money'] = retirement_money
        session['employment_type'] = employment_type
        return redirect(url_for('simulate'))
    return render_template('step3.html')

@app.route('/simulate', methods=['GET'])
def simulate():
    try:
        current_age = session['age']
        current_income = session['current_income']
        savings = session['savings']
        retirement_age = session['retirement_age']
        peak_income = session['peak_income']
        retirement_money = session['retirement_money']
        employment_type = session.get('employment_type')
        housing_status = session['housing_status']
        rent = session.get('rent', 0)
        loan_amount = session.get('loan_amount', 0)
        loan_years = session.get('loan_years', 0)
        housing_expenses = session.get('housing_expenses', 0)
        special_expenses = session.get('special_expenses', [])
        special_incomes = session.get('special_incomes', [])
        housing_changes = session.get('housing_changes', [])
        insurances = session.get('insurances', [])
        working_style = session.get('working_style', None)
    except KeyError as e:
        flash(f"必要なデータが不足しています: {e}")
        return redirect(url_for('step1'))
    
    person_ages = list(range(current_age, 91))
    if working_style is not None:
        var_default = "company" if employment_type == "company" else "self"
        while len(working_style) < len(person_ages):
            working_style.append(var_default)
    
    try:
        var_custom = session.get('custom_incomes')
        if var_custom:
            if isinstance(var_custom, str):
                custom_incomes = json.loads(var_custom)
            else:
                custom_incomes = var_custom
            custom_incomes = [int(x) for x in custom_incomes]
            required_length = retirement_age - current_age + 1
            if len(custom_incomes) < required_length:
                if len(custom_incomes) >= 2:
                    diff = custom_incomes[-1] - custom_incomes[-2]
                else:
                    diff = 0
                while len(custom_incomes) < required_length:
                    custom_incomes.append(custom_incomes[-1] + diff)
                session['custom_incomes'] = custom_incomes
        else:
            custom_incomes = None
        
        from plotly import offline as plotly_offline
        import plotly.graph_objs as go
        
        # ---------------------
        # 資産推移グラフ (シミュレーショングラフ)
        # ---------------------
        ideal_living_expense, person_ages, assets, income, living_expenses, other_expenses, \
          tax_social_insurance, other_incomes, housing_expenses_list, insurance_premiums, \
          peak_age, peak_asset = calculate_living_expenses(
            current_age=current_age,
            retirement_age=retirement_age,
            current_income=current_income,
            current_savings=savings,
            peak_income=peak_income,
            rent=rent,
            car_start_age=0,
            car_price=0,
            car_interval=0,
            housing_status=housing_status,
            retirement_money=retirement_money,
            loan_amount=loan_amount,
            loan_years=loan_years,
            housing_expenses=housing_expenses,
            special_expenses=special_expenses,
            special_incomes=special_incomes,
            housing_changes=housing_changes,
            insurances=insurances,
            employment_type=employment_type,
            custom_incomes=custom_incomes,
            working_style=working_style
        )
        
        # 折れ線グラフ hovertemplate を "%{x}歳, %{y:,.0f}万円"
        line_trace = go.Scatter(
            x=person_ages,
            y=assets,
            mode='lines',
            line=dict(color='blue', width=3),
            hovertemplate='%{x}歳, %{y:,.0f}万円<extra></extra>',
            name=''
        )
        # 下部の塗りつぶし(正の部分)
        fill_trace = go.Scatter(
            x=person_ages,
            y=[a if a > 0 else 0 for a in assets],
            fill='tozeroy',
            fillcolor='rgba(173,216,230,0.3)',
            line=dict(color='rgba(0,0,0,0)'),
            hoverinfo='skip',
            showlegend=False
        )
        
        shape_pink = dict(
            type="rect",
            xref="paper",
            yref="y",
            x0=0,
            x1=1,
            y0=min(assets) - 10,
            y1=0,
            fillcolor="lightpink",
            opacity=0.5,
            layer="below",
            line_width=0
        )
        
        layout = go.Layout(
            title='理想生活費シミュレーション',
            width=1300,
            height=700,
            hovermode='closest',
            shapes=[shape_pink],
            plot_bgcolor='white',
            paper_bgcolor='white',
            showlegend=False,
            dragmode=False,
            xaxis=dict(
                title='年齢',
                dtick=1,
                tickangle=-45,
                showgrid=True,
                gridcolor='lightgray',
                zeroline=False,
                mirror=True,
                linecolor='black',
                linewidth=2,
                range=[person_ages[0], person_ages[-1]],
                fixedrange=True
            ),
            yaxis=dict(
                title='資産額 (万円)',
                range=[min(assets) - 10, max(assets) + 10],
                showgrid=True,
                gridcolor='lightgray',
                zeroline=False,
                mirror=True,
                linecolor='black',
                linewidth=2,
                tickformat=',.0f',
                ticksuffix='万円',
                fixedrange=True
            )
        )
        
        config = {
            "displaylogo": False,
            "displayModeBar": True,
            "modeBarButtonsToRemove": [
                "zoom2d", "pan2d", "select2d", "lasso2d",
                "zoomIn2d", "zoomOut2d", "autoScale2d", "resetScale2d",
                "hoverClosestCartesian", "hoverCompareCartesian",
                "toggleSpikelines"
            ]
        }
        
        fig = go.Figure(data=[fill_trace, line_trace], layout=layout)
        graph_div = plotly_offline.plot(
            fig,
            output_type='div',
            include_plotlyjs=True,
            config=config
        )
        
        # --------------------------------------
        # 年度別支出グラフ (棒グラフ + 収入合計の線)
        # --------------------------------------
        bar_tax = go.Bar(
            x=person_ages,
            y=tax_social_insurance,
            name='税金社会保険料',
            marker=dict(color='tomato'),
            hovertemplate='%{x}歳, %{y:,.0f}万円<extra></extra>'
        )
        bar_living = go.Bar(
            x=person_ages,
            y=living_expenses,
            name='生活費等',
            marker=dict(color='orange'),
            hovertemplate='%{x}歳, %{y:,.0f}万円<extra></extra>'
        )
        bar_housing = go.Bar(
            x=person_ages,
            y=housing_expenses_list,
            name='住居費',
            marker=dict(color='green'),
            hovertemplate='%{x}歳, %{y:,.0f}万円<extra></extra>'
        )
        bar_insurance = go.Bar(
            x=person_ages,
            y=insurance_premiums,
            name='保険料',
            marker=dict(color='blue'),
            hovertemplate='%{x}歳, %{y:,.0f}万円<extra></extra>'
        )
        bar_other = go.Bar(
            x=person_ages,
            y=other_expenses,
            name='その他支出',
            marker=dict(color='purple'),
            hovertemplate='%{x}歳, %{y:,.0f}万円<extra></extra>'
        )
        
        total_income = [income[i] + other_incomes[i] for i in range(len(person_ages))]
        line_income = go.Scatter(
            x=person_ages,
            y=total_income,
            mode='lines',
            line=dict(color='red', width=3),
            name='収入合計',
            yaxis='y',
            hovertemplate='%{x}歳, %{y:,.0f}万円<extra></extra>'
        )
        
        layout_bar = go.Layout(
            title='年度別収支グラフ',
            width=1300,
            height=500,
            barmode='stack',
            legend=dict(
                orientation='h',
                x=0,
                y=1.1
            ),
            hovermode='closest',
            plot_bgcolor='white',
            paper_bgcolor='white',
            dragmode=False,
            xaxis=dict(
                title='年齢',
                dtick=1,
                tickangle=-45,
                showgrid=True,
                gridcolor='lightgray',
                zeroline=False,
                mirror=True,
                linecolor='black',
                linewidth=2,
                range=[person_ages[0], person_ages[-1]],
                fixedrange=True
            ),
            yaxis=dict(
                title='金額 (万円)',
                showgrid=True,
                gridcolor='lightgray',
                zeroline=False,
                mirror=True,
                linecolor='black',
                linewidth=2,
                tickformat=',.0f',
                ticksuffix='万円',
                fixedrange=True
            )
        )
        
        fig_bar = go.Figure(data=[bar_tax, bar_living, bar_housing, bar_insurance, bar_other, line_income],
                            layout=layout_bar)
        
        config_bar = {
            "displaylogo": False,
            "displayModeBar": True,
            "modeBarButtonsToRemove": [
                "zoom2d", "pan2d", "select2d", "lasso2d",
                "zoomIn2d", "zoomOut2d", "autoScale2d", "resetScale2d",
                "hoverClosestCartesian", "hoverCompareCartesian",
                "toggleSpikelines"
            ]
        }
        
        bar_div = plotly_offline.plot(
            fig_bar,
            output_type='div',
            include_plotlyjs=True,
            config=config_bar
        )
        
        # --------------------------------------
        # Excel 出力部分（年齢が横に並ぶ形式に変更）
        # --------------------------------------
        import io
        import base64
        from openpyxl.styles import PatternFill, Font

        # 年間手取りを計算
        income_sum = [income[i] + other_incomes[i] for i in range(len(person_ages))]
        annual_take_home = [income_sum[i] - tax_social_insurance[i] for i in range(len(person_ages))]
        df_data = {
            '年齢': person_ages,
            '収入': income,
            'その他収入': other_incomes,
            '収入合計': income_sum,
            '税金社会保険料': tax_social_insurance,
            '年間手取り': annual_take_home,
            '生活費等': living_expenses,
            '住居費': housing_expenses_list,
            '保険料': insurance_premiums,
            'その他支出': other_expenses,
            '支出合計': [
                tax_social_insurance[i] + living_expenses[i] + housing_expenses_list[i] + insurance_premiums[i] + other_expenses[i]
                for i in range(len(person_ages))
            ],
            '年間貯蓄': [
                income_sum[i] - (tax_social_insurance[i] + living_expenses[i] + housing_expenses_list[i] + insurance_premiums[i] + other_expenses[i])
                for i in range(len(person_ages))
            ],
            '資産額': assets
        }
        df = pd.DataFrame(df_data)
        df.set_index('年齢', inplace=True)
        df = df.T
        df.columns = [f"{col}歳" for col in df.columns]

        # 関数：数値をカンマ付きかつ「万円」を付与してフォーマット
        def format_value(x):
            try:
                x_int = int(round(float(x)))
                return f"{format(x_int, ',')}万円"
            except Exception:
                return x

        for col in df.columns:
            for row in df.index:
                df.at[row, col] = format_value(df.at[row, col])

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Sheet1', index=True)
            worksheet = writer.sheets['Sheet1']
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = max(length + 2, 15)
            # 収入合計：青、年間手取り：薄い青、支出合計：ピンク
            # 年間貯蓄：イエロー、資産額：グリーン
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                if row[0].value == "収入合計":
                    for cell in row:
                        cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                elif row[0].value == "年間手取り":
                    for cell in row:
                        cell.fill = PatternFill(start_color="D0EFFF", end_color="D0EFFF", fill_type="solid")
                elif row[0].value == "支出合計":
                    for cell in row:
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                elif row[0].value == "年間貯蓄":
                    for cell in row:
                        cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
                        try:
                            num = int(str(cell.value).replace("万円", "").replace(",", "").strip())
                            if num <= 0:
                                cell.font = Font(color="FF0000")
                        except Exception:
                            pass
                elif row[0].value == "資産額":
                    for cell in row:
                        cell.fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
                        try:
                            num = int(str(cell.value).replace("万円", "").replace(",", "").strip())
                            if num <= 0:
                                cell.font = Font(color="FF0000")
                        except Exception:
                            pass
        output.seek(0)
        excel_data = base64.b64encode(output.read()).decode()
        
        return render_template(
            'result.html',
            ideal_living_expense=ideal_living_expense,
            graph_div=graph_div,
            bar_div=bar_div,
            special_expenses=special_expenses,
            special_incomes=special_incomes,
            housing_changes=housing_changes,
            insurances=insurances,
            excel_data=excel_data,
            peak_age=peak_age,
            peak_asset=peak_asset
        )
    except Exception as e:
        flash(f"シミュレーションの実行中にエラーが発生しました: {e}")
        return redirect(url_for('step1'))

@app.route('/income_detail', methods=['GET', 'POST'])
def income_detail():
    if request.method == 'POST':
        try:
            employment_type = request.form['employment_type']
            if employment_type == "company":
                current_income = int(request.form['company_income'])
                retirement_age = int(request.form['company_retirement_age'])
                peak_income = int(request.form['company_peak_income'])
                retirement_money_str = request.form.get('retirement_money', None)
                if retirement_money_str:
                    try:
                        retirement_money = json.loads(retirement_money_str)
                    except Exception:
                        retirement_money = session.get('retirement_money', [])
                else:
                    retirement_money = session.get('retirement_money', [])
            else:
                current_income = int(request.form['self_income'])
                retirement_age = int(request.form['self_retirement_age'])
                peak_income = int(request.form['self_peak_income'])
                retirement_money_str = request.form.get('retirement_money', None)
                if retirement_money_str:
                    try:
                        retirement_money = json.loads(retirement_money_str)
                    except Exception:
                        retirement_money = session.get('retirement_money', [])
                else:
                    retirement_money = session.get('retirement_money', [])
        except ValueError:
            flash("収入情報に誤りがあります。")
            return redirect(url_for('step3'))
        if not (current_income > 0):
            flash("現在の年収（または年間手取り）は正の値で入力してください。")
            return redirect(url_for('step3'))
        if not (retirement_age > session['age'] and retirement_age <= 100):
            flash("退職年齢（または引退年齢）は現在の年齢より高く、100歳以下で入力してください。")
            return redirect(url_for('step3'))
        session['current_income'] = current_income
        session['retirement_age'] = retirement_age
        session['peak_income'] = peak_income
        session['retirement_money'] = retirement_money
        session['employment_type'] = employment_type
    else:
        if 'retirement_money' not in session:
            session['retirement_money'] = []
    try:
        current_age = session['age']
        current_income = session['current_income']
        retirement_age = session['retirement_age']
        peak_income = session['peak_income']
    except KeyError as e:
        flash(f"必要な収入情報が不足しています: {e}")
        return redirect(url_for('step3'))
    
    var_ages = list(range(current_age, retirement_age + 1))
    if 'custom_incomes' in session:
        custom_incomes = session['custom_incomes']
        required_length = retirement_age - current_age + 1
        if len(custom_incomes) < required_length:
            if len(custom_incomes) >= 2:
                diff = custom_incomes[-1] - custom_incomes[-2]
            else:
                diff = 0
            while len(custom_incomes) < required_length:
                custom_incomes.append(custom_incomes[-1] + diff)
            session['custom_incomes'] = custom_incomes
        var_incomes = custom_incomes
    else:
        var_incomes = [
            current_income + (peak_income - current_income) * (age - current_age) / (retirement_age - current_age)
            for age in var_ages
        ]
    return render_template("detail.html", ages=var_ages, incomes=var_incomes)

@app.route('/income_detail_apply', methods=['POST'])
def income_detail_apply():
    custom_incomes_json = request.form.get('custom_incomes')
    var_working_style = request.form.get('working_style')
    if var_working_style:
        try:
            session['working_style'] = json.loads(var_working_style)
        except Exception:
            session['working_style'] = None
    if not custom_incomes_json:
        flash("収入データが送信されませんでした。")
        return redirect(url_for('income_detail'))
    try:
        custom_incomes = json.loads(custom_incomes_json)
        if not custom_incomes or len(custom_incomes) == 0:
            flash("収入データが不正です。")
            return redirect(url_for('income_detail'))
        session['current_income'] = int(custom_incomes[0])
        session['peak_income'] = max(custom_incomes)
        session['custom_incomes'] = custom_incomes
    except Exception as e:
        flash(f"収入データの処理中にエラーが発生しました: {e}")
        return redirect(url_for('income_detail'))
    flash("詳細の収入変更が反映されました。")
    return_param = request.form.get('return') or request.args.get('return')
    if return_param == 'edit':
        return redirect(url_for('edit'))
    return redirect(url_for('step3'))

@app.route('/manage/<data_type>', methods=['GET', 'POST'])
@app.route('/manage/<data_type>/<int:index>', methods=['GET', 'POST'])
def manage(data_type, index=None):
    valid_data_types = ['insurance', 'special_expense', 'special_income', 'housing_change']
    if data_type not in valid_data_types:
        flash("無効なデータタイプです。")
        return redirect(url_for('simulate'))
    if request.method == 'POST':
        if data_type == 'insurance':
            return handle_insurance(data_type, index)
        elif data_type == 'special_expense':
            return handle_special_expense(data_type, index)
        elif data_type == 'special_income':
            return handle_special_income(data_type, index)
        elif data_type == 'housing_change':
            return handle_housing_change(data_type, index)
    else:
        data = None
        current_age = session.get('age', 0)
        if index is not None:
            if data_type == 'insurance':
                insurances = session.get('insurances', [])
                if 0 <= index < len(insurances):
                    data = insurances[index]
                else:
                    flash("無効なインデックスです。")
                    return redirect(url_for('simulate'))
            elif data_type == 'special_expense':
                special_expenses = session.get('special_expenses', [])
                if 0 <= index < len(special_expenses):
                    data = special_expenses[index]
                else:
                    flash("無効なインデックスです。")
                    return redirect(url_for('simulate'))
            elif data_type == 'special_income':
                special_incomes = session.get('special_incomes', [])
                if 0 <= index < len(special_incomes):
                    data = special_incomes[index]
                else:
                    flash("無効なインデックスです. ")
                    return redirect(url_for('simulate'))
            elif data_type == 'housing_change':
                housing_changes = session.get('housing_changes', [])
                if 0 <= index < len(housing_changes):
                    data = housing_changes[index]
                else:
                    flash("無効なインデックスです。")
                    return redirect(url_for('simulate'))
        return render_template('manage.html', data_type=data_type, data=data, index=index, current_age=current_age)

def handle_insurance(data_type, index):
    joining_age = request.form.get('joining_age')
    payment_end_age = request.form.get('payment_end_age')
    premium = request.form.get('premium')
    surrender_type = request.form.get('surrender_type')
    try:
        joining_age = int(joining_age) if joining_age else session['age']
        payment_end_age = int(payment_end_age) if payment_end_age else joining_age
        premium = int(premium) if premium else 0
    except ValueError:
        flash("数値フィールドには正しい数値を入力してください。")
        return redirect(url_for('manage', data_type=data_type, index=index))
    surrender_details = {}
    if surrender_type == 'lump_sum':
        surrender_age = request.form.get('surrender_age')
        surrender_amount = request.form.get('surrender_amount')
        try:
            surrender_age = int(surrender_age)
            surrender_amount = int(surrender_amount)
        except (ValueError, TypeError):
            flash("解約返戻金の年齢と金額は数値で入力してください。")
            return redirect(url_for('manage', data_type=data_type, index=index))
        surrender_details = {
            "surrender_age": surrender_age,
            "surrender_amount": surrender_amount
        }
    elif surrender_type == 'annuity':
        annuity_start_age = request.form.get('annuity_start_age')
        annuity_end_age = request.form.get('annuity_end_age')
        annuity_amount = request.form.get('annuity_amount')
        try:
            annuity_start_age = int(annuity_start_age)
            annuity_end_age = int(annuity_end_age)
            annuity_amount = int(annuity_amount)
        except (ValueError, TypeError):
            flash("年金受取の年齢と金額は数値で入力してください。")
            return redirect(url_for('manage', data_type=data_type, index=index))
        if annuity_start_age > annuity_end_age:
            flash("年金受取の開始年齢は終了年齢以下で入力してください。")
            return redirect(url_for('manage', data_type=data_type, index=index))
        surrender_details = {
            "annuity_start_age": annuity_start_age,
            "annuity_end_age": annuity_end_age,
            "annuity_amount": annuity_amount
        }
    elif surrender_type == 'none':
        surrender_details = {
            "surrender_age": 0,
            "surrender_amount": 0,
            "annuity_start_age": 0,
            "annuity_end_age": 0,
            "annuity_amount": 0
        }
    else:
        flash("有効な解約返戻金形式を選択してください。")
        return redirect(url_for('manage', data_type=data_type, index=index))
    insurance = {
        "joining_age": joining_age,
        "payment_end_age": payment_end_age,
        "premium": premium,
        "surrender_type": surrender_type,
        "surrender_details": surrender_details
    }
    insurances = session.get('insurances', [])
    if index is not None:
        if 0 <= index < len(insurances):
            insurances[index] = insurance
            flash("保険が更新されました。")
        else:
            flash("無効なインデックスです。")
            return redirect(url_for('simulate'))
    else:
        insurances.append(insurance)
        flash("保険が追加されました。")
    session['insurances'] = insurances
    session.modified = True
    return redirect(url_for('simulate'))

def handle_special_expense(data_type, index):
    name = request.form.get('expense_name')
    expense_type = request.form.get('expense_type')
    if expense_type == 'single':
        age = request.form.get('age')
        amount = request.form.get('amount')
        if not name or not age or not amount:
            flash("すべてのフィールドを入力してください。")
            return redirect(url_for('manage', data_type=data_type, index=index))
        try:
            age = int(age)
            amount = int(amount)
        except ValueError:
            flash("年齢と金額は数値で入力してください。")
            return redirect(url_for('manage', data_type=data_type, index=index))
        expense = {
            "name": name,
            "type": "single",
            "age": age,
            "amount": amount
        }
    elif expense_type == 'recurring':
        start_age = request.form.get('start_age')
        end_age = request.form.get('end_age')
        amount = request.form.get('amount_recurring')
        interval = request.form.get('interval')
        if not name or not start_age or not end_age or not amount or not interval:
            flash("すべてのフィールドを入力してください。")
            return redirect(url_for('manage', data_type=data_type, index=index))
        try:
            start_age = int(start_age)
            end_age = int(end_age)
            amount = int(amount)
            interval = int(interval)
        except ValueError:
            flash("年齢、金額、間隔は数値で入力してください。")
            return redirect(url_for('manage', data_type=data_type, index=index))
        if start_age > end_age:
            flash("開始年齢は終了年齢以下で入力してください。")
            return redirect(url_for('manage', data_type=data_type, index=index))
        expense = {
            "name": name,
            "type": "recurring",
            "start_age": start_age,
            "end_age": end_age,
            "amount": amount,
            "interval": interval
        }
    else:
        flash("無効な支出タイプです。")
        return redirect(url_for('manage', data_type=data_type, index=index))
    special_expenses = session.get('special_expenses', [])
    if index is not None:
        if 0 <= index < len(special_expenses):
            special_expenses[index] = expense
            flash("特別支出が更新されました。")
        else:
            flash("無効なインデックスです。")
            return redirect(url_for('simulate'))
    else:
        special_expenses.append(expense)
        flash("特別支出が追加されました。")
    session['special_expenses'] = special_expenses
    session.modified = True
    return redirect(url_for('simulate'))

def handle_special_income(data_type, index):
    name = request.form.get('income_name')
    income_type = request.form.get('income_type')
    if income_type == 'single':
        age = request.form.get('age')
        amount = request.form.get('amount')
        if not name or not age or not amount:
            flash("すべてのフィールドを入力してください。")
            return redirect(url_for('manage', data_type=data_type, index=index))
        try:
            age = int(age)
            amount = int(amount)
        except ValueError:
            flash("年齢と金額は数値で入力してください。")
            return redirect(url_for('manage', data_type=data_type, index=index))
        income = {
            "name": name,
            "type": "single",
            "age": age,
            "amount": amount
        }
    elif income_type == 'recurring':
        start_age = request.form.get('start_age')
        end_age = request.form.get('end_age')
        amount = request.form.get('amount_recurring')
        interval = request.form.get('interval')
        if not name or not start_age or not end_age or not amount or not interval:
            flash("すべてのフィールドを入力してください。")
            return redirect(url_for('manage', data_type=data_type, index=index))
        try:
            start_age = int(start_age)
            end_age = int(end_age)
            amount = int(amount)
            interval = int(interval)
        except ValueError:
            flash("年齢、金額、間隔は数値で入力してください。")
            return redirect(url_for('manage', data_type=data_type, index=index))
        if start_age > end_age:
            flash("開始年齢は終了年齢以下で入力してください。")
            return redirect(url_for('manage', data_type=data_type, index=index))
        income = {
            "name": name,
            "type": "recurring",
            "start_age": start_age,
            "end_age": end_age,
            "amount": amount,
            "interval": interval
        }
    else:
        flash("無効な収入タイプです。")
        return redirect(url_for('manage', data_type=data_type, index=index))
    special_incomes = session.get('special_incomes', [])
    if index is not None:
        if 0 <= index < len(special_incomes):
            special_incomes[index] = income
            flash("特別収入が更新されました。")
        else:
            flash("無効なインデックスです。")
            return redirect(url_for('simulate'))
    else:
        special_incomes.append(income)
        flash("特別収入が追加されました。")
    session['special_incomes'] = special_incomes
    session.modified = True
    return redirect(url_for('simulate'))

def handle_housing_change(data_type, index):
    change_age = request.form.get('change_age')
    housing_type = request.form.get('housing_type')
    try:
        change_age = int(change_age)
    except ValueError:
        flash("年齢は数値で入力してください。")
        return redirect(url_for('manage', data_type=data_type, index=index))
    housing_change = {
        "change_age": change_age,
        "housing_type": housing_type
    }
    if housing_type in ['賃貸', 'その他']:
        annual_cost = request.form.get('annual_cost', 0)
        try:
            annual_cost = int(annual_cost)
        except ValueError:
            flash("年間住居費は数値で入力してください。")
            return redirect(url_for('manage', data_type=data_type, index=index))
        housing_change['annual_cost'] = annual_cost
    elif housing_type == '購入':
        property_price = request.form.get('property_price', 0)
        down_payment = request.form.get('down_payment', 0)
        loan_years = request.form.get('loan_years', 0)
        loan_interest_rate = request.form.get('loan_interest_rate', 0)
        other_annual_cost = request.form.get('other_annual_cost', 0)
        try:
            property_price = int(property_price)
            down_payment = int(down_payment)
            loan_years = int(loan_years)
            loan_interest_rate = float(loan_interest_rate)
            other_annual_cost = int(other_annual_cost)
        except ValueError:
            flash("物件価格、頭金、ローン年数、ローン金利、その他年間費用は数値で入力してください。")
            return redirect(url_for('manage', data_type=data_type, index=index))
        housing_change['property_price'] = property_price
        housing_change['down_payment'] = down_payment
        housing_change['loan_years'] = loan_years
        housing_change['loan_interest_rate'] = loan_interest_rate
        housing_change['other_annual_cost'] = other_annual_cost
    else:
        flash("無効な住宅タイプです。")
        return redirect(url_for('manage', data_type=data_type, index=index))
    selling_profit = request.form.get('selling_profit', 0)
    misc_costs = request.form.get('misc_costs', 0)
    try:
        selling_profit = int(selling_profit) if selling_profit else 0
        misc_costs = int(misc_costs) if misc_costs else 0
    except ValueError:
        flash("売却益と諸費用は数値で入力してください。")
        return redirect(url_for('manage', data_type=data_type, index=index))
    housing_change['selling_profit'] = selling_profit
    housing_change['misc_costs'] = misc_costs
    housing_changes = session.get('housing_changes', [])
    if index is not None:
        if 0 <= index < len(housing_changes):
            housing_changes[index] = housing_change
            flash("住宅変更が更新されました。")
        else:
            flash("無効なインデックスです。")
            return redirect(url_for('simulate'))
    else:
        housing_changes.append(housing_change)
        flash("住宅変更が追加されました。")
    session['housing_changes'] = housing_changes
    session.modified = True
    return redirect(url_for('simulate'))

@app.route('/delete/<data_type>/<int:index>', methods=['POST'])
def delete(data_type, index):
    valid_data_types = ['insurance', 'special_expense', 'special_income', 'housing_change']
    if data_type not in valid_data_types:
        flash("無効なデータタイプです。")
        return redirect(url_for('simulate'))
    if data_type == 'insurance':
        insurances = session.get('insurances', [])
        if 0 <= index < len(insurances):
            del insurances[index]
            session['insurances'] = insurances
            session.modified = True
            flash("保険が削除されました。")
        else:
            flash("無効なインデックスです。")
    elif data_type == 'special_expense':
        special_expenses = session.get('special_expenses', [])
        if 0 <= index < len(special_expenses):
            del special_expenses[index]
            session['special_expenses'] = special_expenses
            session.modified = True
            flash("特別支出が削除されました。")
        else:
            flash("無効なインデックスです。")
    elif data_type == 'special_income':
        special_incomes = session.get('special_incomes', [])
        if 0 <= index < len(special_incomes):
            del special_incomes[index]
            session['special_incomes'] = special_incomes
            session.modified = True
            flash("特別収入が削除されました。")
        else:
            flash("無効なインデックスです。")
    elif data_type == 'housing_change':
        housing_changes = session.get('housing_changes', [])
        if 0 <= index < len(housing_changes):
            del housing_changes[index]
            session['housing_changes'] = housing_changes
            session.modified = True
            flash("住宅変更が削除されました。")
        else:
            flash("無効なインデックスです。")
    return redirect(url_for('simulate'))

@app.route('/edit', methods=['GET', 'POST'])
def edit():
    if request.method == 'POST':
        try:
            # 貯蓄などの基本情報
            if 'savings' in request.form:
                try:
                    savings = int(request.form['savings'])
                    if savings < 0:
                        flash("現在の貯蓄額は0以上で入力してください。")
                        return redirect(url_for('edit'))
                    session['savings'] = savings
                except ValueError:
                    flash("すべてのフィールドに正しい数値を入力してください。")
                    return redirect(url_for('edit'))
            housing_status = request.form.get('housing_status')
            rent = request.form.get('rent', 0)
            loan_amount = request.form.get('loan_amount', 0)
            loan_years = request.form.get('loan_years', 0)
            housing_expenses = request.form.get('housing_expenses', 0)
            try:
                rent = int(rent) if rent else 0
                loan_amount = int(loan_amount) if loan_amount else 0
                loan_years = int(loan_years) if loan_years else 0
                housing_expenses = int(housing_expenses) if housing_expenses else 0
            except ValueError:
                flash("すべてのフィールドに正しい数値を入力してください。")
                return redirect(url_for('edit'))
            session['housing_status'] = housing_status
            session['rent'] = rent
            session['loan_amount'] = loan_amount
            session['loan_years'] = loan_years
            session['housing_expenses'] = housing_expenses
            employment_type = request.form.get('employment_type')
            if employment_type:
                session['employment_type'] = employment_type
            
            # 退職金情報（あり／なしと、年齢・金額ペアを取得）
            retirement_option = request.form.get('edit_retirement_option', 'none')
            if retirement_option == 'yes':
                ages_list = request.form.getlist('edit_retirement_age[]')
                amounts_list = request.form.getlist('edit_retirement_amount[]')
                new_retirement_money = []
                for age_str, amt_str in zip(ages_list, amounts_list):
                    age_str = age_str.strip()
                    amt_str = amt_str.strip()
                    if age_str and amt_str:
                        try:
                            age_val = int(age_str)
                            amt_val = int(amt_str)
                            new_retirement_money.append({"age": age_val, "amount": amt_val})
                        except ValueError:
                            pass
                session['retirement_money'] = new_retirement_money
            else:
                session['retirement_money'] = []
            
            # 収入情報
            if employment_type == "company":
                try:
                    current_income = int(request.form['company_income'])
                    retirement_age = int(request.form['company_retirement_age'])
                    peak_income = int(request.form['company_peak_income'])
                except ValueError:
                    flash("すべてのフィールドに正しい数値を入力してください。")
                    return redirect(url_for('edit'))
            else:
                try:
                    current_income = int(request.form['self_income'])
                    retirement_age = int(request.form['self_retirement_age'])
                    peak_income = int(request.form['self_peak_income'])
                except ValueError:
                    flash("すべてのフィールドに正しい数値を入力してください。")
                    return redirect(url_for('edit'))
            if not (current_income > 0):
                flash("現在の年収（または年間手取り）は正の値で入力してください。")
                return redirect(url_for('edit'))
            if not (retirement_age > session['age'] and retirement_age <= 100):
                flash("退職年齢（または引退年齢）は現在の年齢より高く、100歳以下で入力してください。")
                return redirect(url_for('edit'))
            session['current_income'] = current_income
            session['retirement_age'] = retirement_age
            if 'custom_incomes' in session:
                session['peak_income'] = max(session['custom_incomes'])
            else:
                session['peak_income'] = peak_income
            
            flash("編集内容が反映されました。")
            return redirect(url_for('edit'))
        except Exception as e:
            flash(f"編集の反映中にエラーが発生しました: {e}")
            return redirect(url_for('edit'))
    return render_template('edit.html', session=session)

def open_browser():
    webbrowser.open_new("http://localhost:5053/step1")

# ★★★ 新規追加 ★★★
@app.route('/detailed_breakdown')
def detailed_breakdown():
    try:
        current_age = session['age']
        current_income = session['current_income']
        savings = session['savings']
        retirement_age = session['retirement_age']
        peak_income = session['peak_income']
        retirement_money = session.get('retirement_money', [])
        employment_type = session.get('employment_type')
        housing_status = session.get('housing_status')
        rent = session.get('rent', 0)
        loan_amount = session.get('loan_amount', 0)
        loan_years = session.get('loan_years', 0)
        housing_expenses = session.get('housing_expenses', 0)
        special_expenses = session.get('special_expenses', [])
        special_incomes = session.get('special_incomes', [])
        housing_changes = session.get('housing_changes', [])
        insurances = session.get('insurances', [])
        working_style = session.get('working_style', None)
        custom_incomes = session.get('custom_incomes', None)
        
        ideal_living_expense, person_ages, assets, income, living_expenses, other_expenses, \
            tax_social_insurance, other_incomes, housing_expenses_list, insurance_premiums, \
            peak_age, peak_asset = calculate_living_expenses(
                current_age=current_age,
                retirement_age=retirement_age,
                current_income=current_income,
                current_savings=savings,
                peak_income=peak_income,
                rent=rent,
                car_start_age=0,
                car_price=0,
                car_interval=0,
                housing_status=housing_status,
                retirement_money=retirement_money,
                loan_amount=loan_amount,
                loan_years=loan_years,
                housing_expenses=housing_expenses,
                special_expenses=special_expenses,
                special_incomes=special_incomes,
                housing_changes=housing_changes,
                insurances=insurances,
                employment_type=employment_type,
                custom_incomes=custom_incomes,
                working_style=working_style
            )
        
        import pandas as pd
        # 新規：年間手取りを計算
        income_sum = [income[i] + other_incomes[i] for i in range(len(person_ages))]
        annual_take_home = [income_sum[i] - tax_social_insurance[i] for i in range(len(person_ages))]
        df_data = {
            '年齢': person_ages,
            '収入': income,
            'その他収入': other_incomes,
            '収入合計': income_sum,
            '税金社会保険料': tax_social_insurance,
            '年間手取り': annual_take_home,
            '生活費等': living_expenses,
            '住居費': housing_expenses_list,
            '保険料': insurance_premiums,
            'その他支出': other_expenses,
            '支出合計': [
                tax_social_insurance[i] + living_expenses[i] + housing_expenses_list[i] + insurance_premiums[i] + other_expenses[i]
                for i in range(len(person_ages))
            ],
            '年間貯蓄': [
                income_sum[i] - (tax_social_insurance[i] + living_expenses[i] + housing_expenses_list[i] + insurance_premiums[i] + other_expenses[i])
                for i in range(len(person_ages))
            ],
            '資産額': assets
        }
        df = pd.DataFrame(df_data)
        df.set_index('年齢', inplace=True)
        df = df.T
        df.columns = [f"{col}歳" for col in df.columns]
        
        # 関数：数値をカンマ付きかつ「万円」を付与してフォーマット
        def format_value(x):
            try:
                x_int = int(round(float(x)))
                return f"{format(x_int, ',')}万円"
            except Exception:
                return x
        
        for col in df.columns:
            for row in df.index:
                df.at[row, col] = format_value(df.at[row, col])
        
        df_dict = df.to_dict(orient="split")
        return render_template('detailed_breakdown.html', df_dict=df_dict)
    except Exception as e:
        flash(f"詳細収支情報の取得中にエラーが発生しました: {e}")
        return redirect(url_for('simulate'))
# ★★★ 以上 新規追加 ★★★

if __name__ == '__main__':
    from werkzeug.middleware.proxy_fix import ProxyFix
    app.wsgi_app = ProxyFix(app.wsgi_app)
    Timer(1, open_browser).start()
    app.run(debug=False, use_reloader=False, host='0.0.0.0', port=5053)
